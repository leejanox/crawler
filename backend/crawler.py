import threading
import queue
import datetime
import time
import os
import openpyxl as op
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By


# urls 
momBeBe = 'https://cafe.naver.com/f-e/cafes/29434212/menus/0?page=1&size=50' #page=1&size=50
cook82 = 'https://www.82cook.com/entiz/enti.php?bn=15'
momsHolic = 'https://cafe.naver.com/f-e/cafes/10094499/menus/0?page=1&size=50'
lemonT = 'https://cafe.naver.com/f-e/cafes/10298136/menus/0?page=1&size=50'
powderRoom = 'https://cafe.naver.com/f-e/cafes/10050813/menus/0?page=1&size=50'

urls = [momBeBe, cook82, momsHolic, lemonT, powderRoom]

# 큐 
_data_queue = queue.Queue() 
filtered_data_queue_excel = queue.Queue()
filtered_data_queue_csv = queue.Queue()

# 데이터 수집 스레드
class DataCollector(threading.Thread):
    # 생성자
    def __init__(self,url, output_queue):
        super().__init__()
        self.output_queue = output_queue
        self.url = url
        self.driver = webdriver.Chrome()
    
    # 데이터 수집
    def run(self):
        while True:
            try:
                self.Open_driver(self.url)
            except Exception as e:
                print(f'{self.url} 오류 발생 : {e}')
                time.sleep(200) # 200초 대기
    
    def get_data_cook82(self):
        data_list = []
        content = self.driver.find_element(By.XPATH,'//*[@id="bbs"]')
        table = content.find_element(By.TAG_NAME,'table')
        tbody = table.find_element(By.TAG_NAME,'tbody')
        rows = tbody.find_elements(By.TAG_NAME,'tr')
        for row in rows:
            if 'noticeList' in row.get_attribute('class').split(' '):
                continue
            tds = row.find_elements(By.TAG_NAME,'td')
            board = '자유게시판'
            i = tds[0].text.strip()
            title = tds[1].text.strip() # 마지막에 댓글수 나옴
            writer = tds[2].text.strip()
            date = tds[3].text.strip()
            view = tds[4].text.strip()
            row_data = [board, i, title, writer, date, view]
            data_list.append(row_data)
        return data_list
    
    def get_data_naver_cafe(self):
        data_list = []
        content = self.driver.find_element(By.XPATH,'//*[@id="cafe_content"]')
        table = content.find_element(By.TAG_NAME,'table')
        tbodys = table.find_elements(By.TAG_NAME,'tbody')
        for tbody in tbodys:
            if 'board-notice' in tbody.get_attribute('class').split(' '):
                continue
            valid_tbody = tbody
        rows = valid_tbody.find_elements(By.TAG_NAME,'tr')
        for idx, row in enumerate(rows):
            board = row.find_element(By.CLASS_NAME,'board_name').text.strip()
            i = idx + 1
            title = row.find_element(By.CLASS_NAME,'article').text.strip()
            writer = row.find_element(By.CLASS_NAME,'nickname').text.strip()
            date = row.find_element(By.CLASS_NAME,'td_normal.type_date').text.strip()
            view = row.find_element(By.CLASS_NAME,'td_normal.type_readCount').text.strip()
            row_data = [board, i, title, writer, date, view]
            data_list.append(row_data)
        return data_list
            
    # 브라우저 열기
    def Open_driver(self, url):
        data_list = {
            'cook82': [],
            'powderRoom': [],
            'momsHolic': [],
            'lemonT': [],
            'momBeBe': []
        }
        try:
            self.driver.get(url)
            time.sleep(5)
            # 데이터 수집
            try:
                if self.driver.current_url == cook82:
                    data_list['cook82'] = self.get_data_cook82()
                    print(f'data_list_cook82 : {data_list["cook82"]}')
                elif self.driver.current_url == powderRoom:
                    data_list['powderRoom'] = self.get_data_naver_cafe()
                    print(f'data_list_powderRoom : {data_list["powderRoom"]}')
                elif self.driver.current_url == momsHolic:
                    data_list['momsHolic'] = self.get_data_naver_cafe()
                    print(f'data_list_momsHolic : {data_list["momsHolic"]}')
                elif self.driver.current_url == lemonT:
                    data_list['lemonT'] = self.get_data_naver_cafe()
                    print(f'data_list_lemonT : {data_list["lemonT"]}')
                elif self.driver.current_url == momBeBe:
                    data_list['momBeBe'] = self.get_data_naver_cafe()
                    print(f'data_list_momBeBe : {data_list["momBeBe"]}')
                #print(f'{url} 컨텐츠 있음')
            except Exception as e:
                #print(f'{url} 컨텐츠 찾기 실패 : {e}')
                print(e)
        
        except Exception as e:
            # print(f'{url} 열 때 오류 발생 : {e}')
            print(e)
        finally:
            #self.driver.quit()
            data_list = {
                'cook82': data_list['cook82'],
                'powderRoom': data_list['powderRoom'],
                'momsHolic': data_list['momsHolic'],
                'lemonT': data_list['lemonT'],
                'momBeBe': data_list['momBeBe']
            }
            self.output_queue.put(data_list)
            # print('데이터 큐에 저장 완료')

class DataFilter(threading.Thread):
    def __init__(self, input_queue, output_queue_excel, output_queue_csv):
        super().__init__()
        self.input_queue = input_queue
        self.output_queue_excel = output_queue_excel
        self.output_queue_csv = output_queue_csv
        self.prev_data = []
        
    def run(self):
        while True:
            data = self.input_queue.get()
            filtered_data = {
                'cook82': [row for row in data['cook82'] if row not in self.prev_data],
                'powderRoom': [row for row in data['powderRoom'] if row not in self.prev_data],
                'momsHolic': [row for row in data['momsHolic'] if row not in self.prev_data],
                'lemonT': [row for row in data['lemonT'] if row not in self.prev_data],
                'momBeBe': [row for row in data['momBeBe'] if row not in self.prev_data]
            }
            if filtered_data:
                for rows in filtered_data.values():
                    self.prev_data.extend(rows)
                self.output_queue_excel.put(filtered_data)
                self.output_queue_csv.put(filtered_data)
                print(f'필터링 완료 {sum(len(rows) for rows in filtered_data.values())}개 새로운 데이터')
            else:
                print('필터링 실패 or 중복 데이터 없음')
                time.sleep(5)
            self.input_queue.task_done()
            
class DataSaveExcel(threading.Thread):
    def __init__(self, input_queue):
        super().__init__()
        self.input_queue = input_queue
        self.wb = None
        self.curr_date = None
        self.file_path = None
        
    def _create_new_workbook(self, today):
        self.file_path = fr'D:\Kim goeun\database\cafe_croll_{today}.xlsx'
        self.wb = op.Workbook()
        
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        header_row = ['게시판', '번호', '제목', '작성자', '작성일', '조회수']
        sheet_names = ['cook82', 'powderRoom', 'momsHolic', 'lemonT', 'momBeBe']
        
        for sheet_name in sheet_names:
            ws = self.wb.create_sheet(title=sheet_name)
            ws.append(header_row)

        self.wb.save(self.file_path)
        print(f'{today} 엑셀 파일 생성 완료')
        
    def run(self):
        while True:
            data = self.input_queue.get()
            # 날짜 갱신
            today = datetime.datetime.today().strftime('%Y%m%d')
            file_path = fr'D:\Kim goeun\database\cafe_croll_{today}.xlsx'

            if not os.path.exists(file_path):
                self._create_new_workbook(today)
            else:
                self.wb = op.load_workbook(self.file_path)
                print(f'{today} 엑셀 파일 로드 완료')

            for sheet_name in data.keys():
                if sheet_name not in self.wb.sheetnames:
                    self.wb.create_sheet(title=sheet_name)
                    ws = self.wb[sheet_name]
                    ws.append(['게시판', '번호', '제목', '작성자', '작성일', '조회수'])
                else:
                    ws = self.wb[sheet_name]
                for row in data[sheet_name]:
                    ws.append(row)
            self.wb.save(self.file_path)
            self.wb.close()
            self.input_queue.task_done()
            time.sleep(5)

class DataSaveCSV(threading.Thread):
    def __init__(self, input_queue):
        super().__init__()
        self.input_queue = input_queue

    def run(self):
        while True:
            data = self.input_queue.get()
            
            today = datetime.datetime.today().strftime('%Y%m%d')
            folder_path = fr'D:\Kim goeun\database\cafe_croll_{today}'
            os.makedirs(folder_path, exist_ok=True)

            for cafe, rows in data.items():
                file_path = os.path.join(folder_path, f'{cafe}.csv')

                is_new_file = not os.path.exists(file_path)
                with open(file_path, mode='a', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    if is_new_file:
                        writer.writerow(['게시판', '번호', '제목', '작성자', '작성일', '조회수'])
                    writer.writerows(rows)
            
            print(f'{today} CSV 저장 완료')
            self.input_queue.task_done()
            time.sleep(5)
    
def main():
    collectors = []
    for url in urls:
        collector = DataCollector(url, _data_queue)
        collector.start()
        collectors.append(collector)
        
    data_filter = DataFilter(_data_queue, filtered_data_queue_excel, filtered_data_queue_csv)
    data_filter.start()
    
    data_saver = DataSaveExcel(filtered_data_queue_excel)
    data_saver.start()
    
    data_saver_csv = DataSaveCSV(filtered_data_queue_csv)
    data_saver_csv.start()
    
    for collector in collectors:
        collector.join()
    
    data_filter.join()
    data_saver.join()
    data_saver_csv.join()
    
if __name__ == '__main__':
    main()
